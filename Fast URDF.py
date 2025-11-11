import sys
from os import path, listdir, makedirs, remove, rename
from shutil import rmtree
from pandas import read_excel
from requests import get
from zipfile import ZipFile
from io import BytesIO
from winreg import OpenKey, HKEY_CURRENT_USER, QueryValueEx
from time import time, sleep
from math import floor, ceil
from selenium.webdriver import Edge
from selenium.webdriver.common.by import By
from selenium.webdriver.edge.service import Service as EdgeService
from selenium.webdriver.edge.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support.expected_conditions import presence_of_element_located
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from xml.etree.ElementTree import parse
from re import search
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from PyQt5.QtCore import QObject, QThread, pyqtSignal
from PyQt5.QtWidgets import QApplication, QFileDialog, QWidget, QVBoxLayout, QHBoxLayout, QTextEdit, QLineEdit, QPushButton, QSizePolicy
from PyQt5.QtGui import QIcon

# 版本号
__version__ = '1.3'

# 报错及提示输出包装函数

# 报错提示
def print_ten_pentagram():
    print('★' * 10)
    return

# 无效
def print_error_void(data_type):
    print(f" 【Error】 {data_type}无效，请检查！")
    return

# 找文件、找到文件、未找到文件、读取文件
def print_info_file_finding(folder_name, file_type):
    print(f" --> 正在【{folder_name}】中搜索 {file_type} 文件 ... 请勿中止！")
    return

def print_info_file_found(file_name, file_type):
    print(f" --> 找到 {file_type} 文件【{file_name}】！")
    return

def print_error_file_not_found(file_type):
    print(f" 【Error】 未找到 {file_type} 文件！")
    return

def print_info_file_reading(file_name, file_type):
    print(f" --> 正在读取 {file_type} 文件:【{file_name}】！")
    return

# 从原始 Excel 文件中读取数据
def print_error_data_miss(*args):
    # args[0] -> link_name
    # args[1] -> data_type
    if len(args) == 0:
        print(" 【Error】 原始 Excel 文件中连杆属性数据缺失，请检查！")
    elif len(args) == 1:
        print(f" 【Error】 原始 Excel 文件中{args[0]} 数据缺失，请检查！")
    elif len(args) == 2:
        print(f" 【Error】 原始 Excel 文件中{args[0]} 的{args[1]}数据缺失，请检查！")
    return

# 将数据写入欲生成的 Excel 文件中
def print_info_excel_written(*args):
    # args[0] -> link_names
    # args[1] -> output_file
    # 或
    # args[0] -> link_name
    # args[1] -> data_type
    # args[2] -> output_file
    if len(args) == 2:
        print(f" --> 共 {len(args[0])} 组连杆数据已写入【{args[1]}】！")
        for i, link_name in enumerate(args[0]):
            print(f" --> Link {i + 1} -- {link_name}")

    elif len(args) == 3:
        print(f" --> {args[0]} 的{args[1]}已写入【{args[2]}】！")
    return

def print_info_excel_finish(excel_file):
    print(f" --> Excel 文件已完整构建！")
    print(f" --> 重要提示！经整理的连杆属性文件已保存至【{excel_file}】！")
    print(f" --> 重要提示！请勿随意增删改工作目录文件夹结构！")
    return

def print_info_excel_unfinished():
    print(" --> 重要提示！工作目录下的 excel 文件夹内的连杆属性文件不完整！请重新生成！")
    return

def print_error_excel_unfinished():
    print(" 【Error】 Excel 文件未完整构建！")
    return

# 使用生成的 Excel 文件中的数据（读取 + 替换）
def print_info_excel_reading(*args):
    # args[0] -> link_name
    # 或
    # args[0] -> link_name
    # args[1] -> state
    # 或
    # args[0] -> link_name
    # args[1] -> state
    # args[2] -> link_data
    if len(args) == 1:
        print(f" --> {args[0]} 数据替换成功！")
    if len(args) == 2:
        print(f" --> 在生成的 Excel 文件中{args[1]}到连杆名称：{args[0]}！")
    elif len(args) == 3:
        print(f" --> 在生成的 Excel 文件中{args[1]}到 {args[0]} 数据：\n{args[2]}")

def print_error_excel_reading(link_name, state):
    print(f" 【Error】 未能在生成的 Excel 文件中{state}到连杆名称：{link_name}！")
    return

# URDF 文件中统一 stl
def print_error_urdf_checking(link_name):
    print(f" 【Error】 {link_name} 的 <mesh> 标签缺失，请检查！")
    return

def print_info_urdf_stl():
    print(" --> 已将 URDF 文件中全部【.STL】扩展名修改为【.stl】！")
    return

# meshes 文件夹中统一 stl
def print_error_meshes_stl(*args):
    # args[0] -> link_name
    if len(args) == 0:
        print(" 【Error】 未找到 meshes 文件夹，请检查！")
    elif len(args) == 1:
        print(f" 【Error】 在 meshes 文件夹中，未找到 {args[0]} 的【.STL】或【.stl】文件，请检查！")
    return

def print_info_meshes_stl(*args):
    # args[0] -> state（删除）
    # args[1] -> old_link_name
    # 或
    # args[0] -> state（重命名）
    # args[1] -> old_link_name
    # args[2] -> new_link_name
    if len(args) == 2:
        print(f" --> 在 meshes 文件夹中，{args[0]}文件：【{args[1]}】成功！")
    elif len(args) == 3:
        print(f" --> 在 meshes 文件夹中，{args[0]}文件：【{args[1]}】-->【{args[2]}】成功！")

# 第一个功能
def print_info_task1(data_name):
    print(f" --> {data_name}修改完成！")
    return

def print_error_task1(data_name):
    print(f" 【Error】 {data_name}修改失败！")
    return

# 检测 Edge 浏览器版本
def print_error_edge_version(*args):
    # args[0] -> e
    if len(args) == 0:
        print(" 【Error】 未能检测到 Edge 浏览器版本！请检查是否已下载 Edge 浏览器！")
    elif len(args) == 1:
        print(f" 【Error】 {args[0]}")
    return

# 准备 Edge 驱动
def print_info_edge_preparing(state):
    print(f" --> 正在{state} Edge 驱动 ... 请勿中止！")
    return

# 存在 Edge 驱动
def print_info_edge_existed(version, msedgedriver_path):
    print(f" --> 已存在适配当前 Edge 浏览器 {version} 版本的【{msedgedriver_path}】，无需重新下载！")
    return

# 下载 Edge 驱动
def print_info_edge_download(download_msedgedriver_path):
    print(f" --> 已下载【msedgedriver.exe】并解压至【{download_msedgedriver_path}】！")
    return

def print_error_edge_download(version):
    print(f" 【Error】 适配当前 Edge 浏览器 {version} 版本的【msedgedriver.exe】下载失败！")
    return

def print_error_edge_download_details(e):
    print(f" 【Error】 下载【msedgedriver.exe】遇到问题：{e}")
    return

# 初始化 Edge 驱动
def print_error_edge_initialization():
    print(" 【Error】 Edge WebDriver 初始化失败！")
    return

# 上传文件
def print_info_file_uploading(file_path):
    print(f" --> 正在上传文件【{file_path}】到 Mesh Simplification 网站 ... 请勿中止！")
    return

# 下载文件
def print_info_file_download(file_path):
    print(f" --> 成功从 Mesh Simplification 网站下载文件到【{file_path}】！")
    return

# 文件大小核对
def print_info_size_detection(size_threshold_mb):
    print(f" --> meshes 文件夹中已无超过 {size_threshold_mb}MB 的【.stl】文件！")
    return

# 第二个功能
def print_info_task2():
    print(" --> 【.stl】文件的简化已完成！")
    return

def print_error_task2():
    print(" 【Error】 【.stl】文件的简化未完成！")
    return

# Fast URDF 完成
def print_info_fast_urdf_finish():
    print(" --> Fast URDF 运行完成！")
    return

# 第〇个功能小函数开始
# find_urdf_file(work_space)
# find_excel_file(work_space)
# sort_mass_properties(work_space, initial_excel_file, urdf_file)

# 自动查找 URDF 文件
def find_urdf_file(work_space):
    
    # 在 urdf 文件夹中查找 URDF 文件
    urdf_folder = path.join(work_space, 'urdf')
    print_info_file_finding(urdf_folder, 'URDF')

    urdf_file = None
    for file in listdir(urdf_folder):
        if file.endswith('.urdf'):
            urdf_file = path.join(urdf_folder, file)
            break

    return urdf_file

# 自动查找 Excel 文件
def find_excel_file(work_space):

    # 在 excel 文件夹中查找 Excel 文件
    excel_folder = path.join(work_space, 'excel')
    print_info_file_finding(excel_folder, 'Excel')

    excel_file = None
    for file in listdir(excel_folder):
        if file.endswith(('.xlsx', '.xls')):
            excel_file = path.join(excel_folder, file)
            break
    
    return excel_file

def sort_mass_properties(work_space, initial_excel_file, urdf_file):
    # 输出文件夹
    output_excel_folder = path.join(work_space, 'excel')
    if path.exists(output_excel_folder):
        # 已存在输出文件夹，则删除输出文件夹及其中所有内容
        rmtree(output_excel_folder)
    # 不存在输出文件夹，则创建输出文件夹
    makedirs(output_excel_folder, exist_ok=True)

    output_excel_file = path.join(output_excel_folder, 'mass_properties_of_links.xlsx')

    # 读取原始 Excel 文件
    df = read_excel(initial_excel_file, header=None)

    unordered_link_names = []
    unordered_link_names_positions = []

    unordered_mass_values = []
    unordered_xyz_values = []
    unordered_inertia_values = []

    # 按列逐一扫描
    for col in range(df.shape[1]):
        for row in range(df.shape[0]):
            cell_value = df.iat[row, col]
            # 检查单元格是否为字符串且以"_cs"结尾
            if isinstance(cell_value, str) and cell_value.endswith('_cs'):
                # 提取空格后面的部分
                if ' ' in cell_value:
                    parts = cell_value.rsplit(' ', 1)
                    if len(parts) == 2:  # 从右侧分割，最多分成两个部分
                        link_name = parts[-1][:-3]  # 取最后一部分，去掉最后的"_cs"
                        unordered_link_names.append(link_name)  # 记录结果
                        unordered_link_names_positions.append((row, col))  # 记录位置
    
    # 提取质量属性
    for pos in unordered_link_names_positions:
        row, col = pos
        mass_numbers = []
        mass_cells = []  # 用于记录所有含有等于号的单元格

        # 向下扫描直到遇到不含等于号的单元格
        for search_row in range(row + 1, df.shape[0]):  # 从当前位置的下一行开始向下扫描
            cell_value = df.iat[search_row, col]
            # 如果是含等于号的单元格，记录该单元格
            if isinstance(cell_value, str) and '=' in cell_value:
                mass_cells.append((search_row, col))  # 记录含有等于号的单元格位置
            # 如果是非空且不含等于号的单元格，停止扫描
            elif isinstance(cell_value, str) and cell_value.strip() != "":
                break  # 遇到非空且不含等于号的格子，停止扫描

        # 如果找到了至少3个含有等于号的格子，从倒数第三个提取质量属性
        if len(mass_cells) >= 3:
            mass_row, mass_col = mass_cells[-3]  # 倒数第三个含等于号的格子
            mass_cell_value = df.iat[mass_row, mass_col]

            # 使用正则表达式提取数字
            # 【[-+]?】
            # 匹配一个可选的符号（ + 或 - ）
            # 【?】表示前面的字符可以出现 0 次或 1 次

            # 【\d*】
            # 匹配数字（0-9）
            # 【*】表示前面的字符可以重复任意次数（包括 0 次）

            # 【\.】
            # 匹配小数点

            # 【\d+】
            # 匹配数字（0-9）
            # 【+】表示前面的字符必须至少出现 1 次

            # 【|】
            # 或运算符，表示可以匹配左侧或右侧的模式
            match = search(r'[-+]?\d*\.\d+|[-+]?\d+', mass_cell_value)
            if match:
                number_part = match.group(0)  # 获取匹配到的数字
                mass_numbers.append(number_part)  # 记录结果

            unordered_mass_values.append(mass_numbers)
    
    # 提取质心属性
    for pos in unordered_link_names_positions:
        row, col = pos
        col += 1  # 往右一格

        # 寻找第一组三个相连的有等于号的格子
        for search_row in range(row + 1, df.shape[0]):
            cell_value = df.iat[search_row, col]
            # 检查单元格是否为字符串且含有【=】
            if isinstance(cell_value, str) and '=' in cell_value:
                # 检查连续的三个格子
                next_cell_value = df.iat[search_row + 1, col] if search_row + 1 < df.shape[0] else None
                next_next_cell_value = df.iat[search_row + 2, col] if search_row + 2 < df.shape[0] else None

                if (isinstance(next_cell_value, str) and '=' in next_cell_value) and (isinstance(next_next_cell_value, str) and '=' in next_next_cell_value):
                    # 提取三个质心属性数字
                    xyz_numbers = []
                    offsets = [(0, 0), (1, 0), (2, 0)]  # 位置偏移量

                    for offset in offsets:
                        xyz_row = search_row + offset[0]
                        xyz_col = col + offset[1]

                        # 确认为质心属性单元格
                        xyz_cell_value = df.iat[xyz_row, xyz_col]
                        # 使用正则表达式提取数字
                        match = search(r'[-+]?\d*\.\d+|[-+]?\d+', xyz_cell_value)
                        if match:
                            number_part = match.group(0)  # 获取匹配到的数字
                            xyz_numbers.append(number_part)  # 记录结果

                    if len(xyz_numbers) == 3:
                        unordered_xyz_values.append(xyz_numbers)
                    break

    # 提取惯量属性
    for pos in unordered_link_names_positions:
        row, col = pos
        col += 1  # 往右一格

        # 寻找第三组三个相连的有等于号的格子
        found_groups = 0
        for search_row in range(row + 1, df.shape[0]):
            cell_value = df.iat[search_row, col]
            if isinstance(cell_value, str) and '=' in cell_value:
                # 检查连续的三个格子
                next_cell_value = df.iat[search_row + 1, col] if search_row + 1 < df.shape[0] else None
                next_next_cell_value = df.iat[search_row + 2, col] if search_row + 2 < df.shape[0] else None

                if (isinstance(next_cell_value, str) and '=' in next_cell_value) and (isinstance(next_next_cell_value, str) and '=' in next_next_cell_value):
                    found_groups += 1
                    if found_groups == 3:  # 找到第三组
                        # 提取六个惯量属性数字
                        inertia_numbers = []
                        offsets = [(0, 0), (1, 0), (2, 0), (1, 1), (2, 1), (2, 2)]  # 位置偏移量

                        for offset in offsets:
                            inertia_row = search_row + offset[0]
                            inertia_col = col + offset[1]

                            # 确认为惯量属性单元格
                            inertia_cell_value = df.iat[inertia_row, inertia_col]
                            match = search(r'[-+]?\d*\.\d+|[-+]?\d+', inertia_cell_value)
                            if match:
                                number_part = match.group(0)  # 获取匹配到的数字
                                inertia_numbers.append(number_part)  # 记录结果

                        if len(inertia_numbers) == 6:
                            unordered_inertia_values.append(inertia_numbers)
                        break
          
    # 创建一个字典，组合 link_names 和对应的属性值
    mass_properties_dict = {}
    
    # 解析 URDF 文件
    tree = parse(urdf_file)
    root = tree.getroot()

    # 获取所有的 link_names
    link_names = []
    for link in root.findall('.//link'):
        link_name = link.get('name')  # 获取 link 的 name 属性
        if link_name:  # 确保 link_name 不为空
            link_names.append(link_name)  # 添加到列表中

    for link_name in link_names:
        if link_name in unordered_link_names:
            index = unordered_link_names.index(link_name)
            mass_properties_dict[link_name] = [
                unordered_mass_values[index] if index < len(unordered_mass_values) else None,
                unordered_xyz_values[index] if index < len(unordered_xyz_values) else None,
                unordered_inertia_values[index] if index < len(unordered_inertia_values) else None
            ]
        else:
            print_ten_pentagram()
            print_error_data_miss(link_name)

    if path.exists(output_excel_file):
        # 删除已存在的文件
        remove(output_excel_file)

    # 创建新的工作簿
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Sheet1'

    worksheet.merge_cells('B3:C3')
    worksheet['B3'] = '质量（千克）'
    worksheet.merge_cells('B4:B6')
    worksheet['B4'] = '质心（米）'
    worksheet.merge_cells('B7:B12')
    worksheet['B7'] = '惯量'
    worksheet['C4'] = 'x'
    worksheet['C5'] = 'y'
    worksheet['C6'] = 'z'
    worksheet['C7'] = 'xx'
    worksheet['C8'] = 'xy'
    worksheet['C9'] = 'xz'
    worksheet['C10'] = 'yy'
    worksheet['C11'] = 'yz'
    worksheet['C12'] = 'zz'

    # 检查标志
    flag = False

    # 遍历字典，输出到 Excel 文件
    row_start = 2  # 从第二行开始输出
    col_start = 4  # 从第四列开始输出

    # 写入连杆名称
    for i, link_name in enumerate(link_names):
        worksheet.cell(row=row_start, column=col_start + i, value=link_name)
        print_info_excel_written(link_name, '连杆名称', output_excel_file)

    # 写入质量属性
    for i, link_name in enumerate(link_names):
        if link_name in mass_properties_dict:
            mass_values = mass_properties_dict[link_name][0]
            if mass_values is not None and len(mass_values) == 1:
                for j in range(1):
                    worksheet.cell(row=row_start + 1 + j, column=col_start + i, value=mass_values[j])
                    print_info_excel_written(link_name, '质量属性', output_excel_file)

            else:
                flag = True
                for j in range(1):
                    worksheet.cell(row=row_start + 1 + j, column=col_start + i, value='NaN')
                    print_ten_pentagram()
                    print_error_data_miss(link_name, '质量属性')

    # 写入质心属性
    for i, link_name in enumerate(link_names):
        if link_name in mass_properties_dict:
            xyz_values = mass_properties_dict[link_name][1]
            if xyz_values is not None and len(xyz_values) == 3:
                for j in range(3):
                    worksheet.cell(row=row_start + 2 + j, column=col_start + i, value=xyz_values[j])
                    print_info_excel_written(link_name, '质心属性', output_excel_file)

            else:
                flag = True
                for j in range(3):
                    worksheet.cell(row=row_start + 2 + j, column=col_start + i, value='NaN')
                    print_ten_pentagram()
                    print_error_data_miss(link_name, '质心属性')

    # 写入惯量属性
    for i, link_name in enumerate(link_names):
        if link_name in mass_properties_dict:
            inertia_values = mass_properties_dict[link_name][2]
            if inertia_values is not None and len(inertia_values) == 6:
                for j in range(6):
                    worksheet.cell(row=row_start + 5 + j, column=col_start + i, value=inertia_values[j])
                    print_info_excel_written(link_name, '惯量属性', output_excel_file)

            else:
                flag = True
                for j in range(6):
                    worksheet.cell(row=row_start + 5 + j, column=col_start + i, value='NaN')
                    print_ten_pentagram()
                    print_error_data_miss(link_name, '惯量属性')
    
    # 表格美化
    # 设置边框样式
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in worksheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                if cell.value:  # 只对有内容的单元格进行处理
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = border
                    if any('\u4e00' <= char <= '\u9fff' for char in cell.value):  # 判断是否为中文
                        cell.font = Font(name='KaiTi')
                    else:
                        cell.font = Font(name='Times New Roman')

    # 添加外边框到合并区域
    for row in worksheet['B3':'C12']:
        for cell in row:
            cell.border = border

    # 保存工作簿
    workbook.save(output_excel_file)
    workbook.close()  # 关闭工作簿

    print_info_excel_written(link_names, output_excel_file)

    if flag:
        print_ten_pentagram()
        print_error_data_miss()
        raise ValueError

# 第〇个功能小函数结束
# find_files(work_space)
# sort_mass_properties(work_space, initial_excel_file, urdf_file)

# 第〇个功能大函数开始
def build_excel_file(work_space, size_threshold_mb, download_msedgedriver_path, initial_excel_file):

    if not isinstance(work_space, str) or not path.isdir(work_space):
        print_ten_pentagram()
        print_error_void('工作目录')
        print_info_excel_unfinished()
        return
    
    if not isinstance(initial_excel_file, str) or not path.isfile(initial_excel_file) or not initial_excel_file.endswith(('.xlsx', '.xls')):
        print_ten_pentagram()
        print_error_void('原始 Excel 文件')
        print_info_excel_unfinished()
        return
    
    # 自动定位 URDF 文件
    try:
        urdf_file = find_urdf_file(work_space)
        print_info_file_found(urdf_file, 'URDF')
    except Exception as e:
        print_ten_pentagram()
        print_error_file_not_found('URDF')
        print_info_excel_unfinished()
        return

    try:
        sort_mass_properties(work_space, initial_excel_file, urdf_file)
        excel_file = find_excel_file(work_space)
        print_info_excel_finish(excel_file)
    except Exception as e:
        print_ten_pentagram()
        print_error_excel_unfinished()
        print_info_excel_unfinished()
        return
            
# 第〇个功能大函数结束

# 第一个功能小函数开始
# find_files(work_space)  # 第〇个功能小函数
# read_excel_data(excel_file)
# modify_urdf_with_excel_data(urdf_file, excel_file)
# modify_urdf_structure(urdf_file)
# modify_mesh_filenames(urdf_file)
# modify_mesh_folder_files(urdf_file, work_space)

# 读取 Excel 文件中的所有列数据
def read_excel_data(excel_file):

    print_info_file_reading(excel_file, 'Excel')

    # 打开 Excel 文件
    workbook = load_workbook(excel_file, data_only=True)
    sheet = workbook.active
    data = {}

    # 从第4列开始遍历 Excel 文件的每列，每列从第2行开始，到第12行结束
    for col in sheet.iter_cols(min_row=2, max_row=12, min_col=4, values_only=True):
        link_name = col[0]  # 第一行是 link name
        print_info_excel_reading(link_name, '读取')
        if link_name:  # 确保 link_name 不为空
            original_data = list(col[1:])  # 后面10行为数据
            print_info_excel_reading(link_name, '读取', original_data)
            data[link_name] = original_data
    
    workbook.close()  # 关闭工作簿
    return data

# 修改 URDF 文件中的 <inertial> 数据
def modify_urdf_with_excel_data(urdf_file, excel_file):

    # 读取 Excel 文件
    excel_data = read_excel_data(excel_file)

    # 解析 URDF 文件
    tree = parse(urdf_file)
    root = tree.getroot()

    # 1、先检索 Excel 文件中是否有全部对应的 link name
    # 遍历 URDF 文件中所有的 <link> 标签
    for link in root.findall('.//link'):
        link_name = link.get('name')
        if link_name in excel_data:  # 如果在 Excel 数据中找到对应的 link name
            print_info_excel_reading(link_name, '匹配')
        else:
            print_ten_pentagram()
            print_error_excel_reading(link_name, '匹配')
            raise ValueError
        
    # 2、再进行 URDF 文件中的数据替换
    # 遍历 URDF 文件中所有的 <link> 标签
    for link in root.findall('.//link'):
        link_name = link.get('name')
        if link_name in excel_data:  # 如果在 Excel 数据中找到对应的 link name
            inertial = link.find('inertial')
            if inertial is not None:
                # 获取 <mass>、<origin>、<inertia>
                mass = inertial.find('mass')
                origin = inertial.find('origin')
                inertia = inertial.find('inertia')

                # 从 excel_data 中读取数据
                new_data = excel_data[link_name]

                mass_value = new_data[0]       # 第1个是 mass
                xyz = new_data[1:4]            # 第2到4个是 xyz
                inertia_values = new_data[4:]  # 剩下的是惯性矩阵

                # 更新 <mass> 标签中的 value
                mass.set('value', str(mass_value))

                # 更新 <origin> 标签中的 xyz
                origin.set('xyz', f"{xyz[0]} {xyz[1]} {xyz[2]}")

                # 更新 <inertia> 标签中的 ixx、ixy、ixz、iyy、iyz、izz
                inertia.set('ixx', str(inertia_values[0]))
                inertia.set('ixy', str(inertia_values[1]))
                inertia.set('ixz', str(inertia_values[2]))
                inertia.set('iyy', str(inertia_values[3]))
                inertia.set('iyz', str(inertia_values[4]))
                inertia.set('izz', str(inertia_values[5]))

                print_info_excel_reading(link_name)

    # 写回 URDF 文件
    tree.write(urdf_file)

# 在 URDF 文件顶部添加 XML 声明，并在 <robot> 标签后添加 <mujoco> 标签
def modify_urdf_structure(urdf_file):
    # 读取 URDF 文件内容
    with open(urdf_file, 'r', encoding='utf-8') as file:
        urdf_content = file.read()

    # 添加 <?xml version="1.0" encoding="utf-8"?>
    if not urdf_content.startswith('<?xml version="1.0"'):
        urdf_content = '<?xml version="1.0" encoding="utf-8"?>\n' + urdf_content

    # 查找 <robot> 标签并在其后添加 <mujoco> 元素，但先检查是否已经存在 <mujoco>
    if '<mujoco>' not in urdf_content:
        robot_tag_index = urdf_content.find('<robot')
        if robot_tag_index != -1:
            closing_bracket_index = urdf_content.find('>', robot_tag_index)
            if closing_bracket_index != -1:
                # 在 <robot> 标签后添加 <mujoco> 元素
                mujoco_text = '''
  <mujoco>
    <compiler meshdir="../meshes/" balanceinertia="true" discardvisual="true" />
  </mujoco>'''
                urdf_content = urdf_content[:closing_bracket_index + 1] + mujoco_text + urdf_content[closing_bracket_index + 1:]

    # 写回 URDF 文件
    with open(urdf_file, 'w', encoding='utf-8') as file:
        file.write(urdf_content)

# 将 URDF 文件中的【.STL】扩展名修改为【.stl】
def modify_mesh_filenames(urdf_file):

    # 解析 URDF 文件
    tree = parse(urdf_file)
    root = tree.getroot()

    # 1、先检查 <mesh> 标签的数量
    # 遍历所有 <link> 标签
    for link in root.findall('.//link'):
        link_name = link.get('name')  # 获取 <link> 的 name 属性
        # 找到 <link> 下的所有 <mesh> 标签
        meshes = link.findall('.//mesh')

        # 如果某个 <link> 中的 <mesh> 标签少于 2 个，报错并终止
        if len(meshes) < 2:
            print_ten_pentagram()
            print_error_urdf_checking(link_name)
            raise ValueError
        
    # 2、再修改 <mesh> 标签中的扩展名
    # 遍历所有 <mesh> 标签
    for mesh in root.findall('.//mesh'):
        filename = mesh.get('filename')
        if filename and filename.endswith('.STL'):
            # 将 .STL 替换为 .stl
            new_filename = filename.replace('.STL', '.stl')
            mesh.set('filename', new_filename)

    # 写回 URDF 文件
    tree.write(urdf_file)
    print_info_urdf_stl()

# 将 meshes 文件夹中的【.STL】扩展名修改为【.stl】
def modify_mesh_folder_files(urdf_file, work_space):

    # 找到 meshes 文件夹
    meshes_folder = path.join(work_space, 'meshes')

    # 检查 meshes 文件夹是否存在
    if not path.exists(meshes_folder):
        print_ten_pentagram()
        print_error_meshes_stl()
        raise FileNotFoundError

    # 获取文件夹中所有文件的名称
    mesh_files = listdir(meshes_folder)

    # 解析 URDF 文件并获取所有连杆名称
    tree = parse(urdf_file)
    root = tree.getroot()
    link_names = [link.get('name') for link in root.findall('.//link')]

    # 遍历 URDF 中的每个连杆名称，检查对应的【.STL】文件
    for link_name in link_names:
        stl_file = f"{link_name}.STL"
        stl_file_lower = f"{link_name}.stl"

        # 如果连杆的【.STL】文件既不以大写也不以小写存在，则报错
        if stl_file not in mesh_files and stl_file_lower not in mesh_files:
            print_ten_pentagram()
            print_error_meshes_stl(link_name)
            raise FileNotFoundError

        # 如果【.STL】文件存在，且是大写，则修改为小写
        if stl_file in mesh_files:
            new_file_name = stl_file.replace('.STL', '.stl')
            rename(path.join(meshes_folder, stl_file), path.join(meshes_folder, new_file_name))
            print_info_meshes_stl('重命名', stl_file, new_file_name)

        # 检查是否有带前缀的文件
        prefixed_file = f"simplify_{link_name}.stl"
        if prefixed_file in mesh_files:
            original_file_path = path.join(meshes_folder, stl_file_lower)
            prefixed_file_path = path.join(meshes_folder, prefixed_file)
            
            # 删除无前缀的文件，重命名带前缀的文件
            if path.exists(original_file_path):
                remove(original_file_path)
                print_info_meshes_stl('删除', stl_file_lower)
                
            rename(prefixed_file_path, original_file_path)
            print_info_meshes_stl('重命名', prefixed_file, stl_file_lower)

# 第一个功能小函数结束
# find_files(work_space)  # 第〇个功能小函数
# read_excel_data(excel_file)
# modify_urdf_with_excel_data(urdf_file, excel_file)
# modify_urdf_structure(urdf_file)
# modify_mesh_filenames(urdf_file)
# modify_mesh_folder_files(urdf_file, work_space)

# 第一个功能大函数 urdf_data_replace 开始
def urdf_data_replace(work_space, size_threshold_mb, download_msedgedriver_path, initial_excel_file):

    if not isinstance(work_space, str) or not path.isdir(work_space):
        print_ten_pentagram()
        print_error_void('工作目录')
        return

    # 自动定位 URDF 文件
    try:
        urdf_file = find_urdf_file(work_space)
        print_info_file_found(urdf_file, 'URDF')
    except Exception as e:
        print_ten_pentagram()
        print_error_file_not_found('URDF')
        return

    # 自动定位 Excel 文件
    try:
        excel_file = find_excel_file(work_space)
        print_info_file_found(excel_file, 'Excel')
    except Exception as e:
        print_ten_pentagram()
        print_error_file_not_found('Excel')
        return
    
    try:
        # 修改 URDF 文件中的 <inertial> 数据
        modify_urdf_with_excel_data(urdf_file, excel_file)

        # 在 URDF 文件顶部添加 XML 声明，并在 <robot> 标签后添加 <mujoco> 标签
        modify_urdf_structure(urdf_file)

        # 将 URDF 文件中的【.STL】扩展名修改为【.stl】
        modify_mesh_filenames(urdf_file)

        print_info_task1('URDF 文件')
    except Exception as e:
        print_ten_pentagram()
        print_error_task1('URDF 文件')
        return

    try:
        # 将 meshes 文件夹中的【.STL】扩展名修改为【.stl】
        modify_mesh_folder_files(urdf_file, work_space)

        print_info_task1('meshes 文件夹')
    except Exception as e:
        print_ten_pentagram()
        print_error_task1('meshes 文件夹')
        return

# 第一个功能大函数 urdf_data_replace 结束

# 第二个功能小函数开始

# 启动浏览器，准备文件
# get_edge_version()
# download_msedgedriver(download_msedgedriver_path)
# setup_driver(download_msedgedriver_path, download_dir)
# find_large_stl_files(meshes_folder_path, size_threshold_mb)

# 上传文件，下载文件
# get_current_slider_position(driver, compression_slider)
# calculate_compression_adjustment(file_size, current_position, size_threshold_mb)
# adjust_compression_slider(driver, compression_slider, file_size, size_threshold_mb)
# wait_for_upload_and_download_buttons(driver)
# wait_for_class_change(driver)
# get_latest_downloaded_file(meshes_folder_path, timeout=30, interval=1)
# upload_and_simplify_stl(driver, file_path, meshes_folder_path, size_threshold_mb)

# 文件重命名
# modify_mesh_folder_files_again(work_space, urdf_file)
# find_files(work_space)  # 第〇个功能小函数

# 自动获取 Edge 浏览器版本
def get_edge_version():
    try:
        key = OpenKey(HKEY_CURRENT_USER, r"Software\Microsoft\Edge\BLBeacon")
        version, _ = QueryValueEx(key, "version")
        return version
    except Exception as e:
        print_ten_pentagram()
        print_error_edge_version(e)
        return None
 
# 自动下载并解压 msedgedriver.exe
def download_msedgedriver(download_msedgedriver_path):

    print_info_edge_preparing('安装')

    if not download_msedgedriver_path:
        print_ten_pentagram()
        print_error_void('Edge 驱动下载目标路径')
        raise ValueError

    if not path.isdir(download_msedgedriver_path):
        makedirs(download_msedgedriver_path)
        
    version = get_edge_version()
    if version is None:
        print_ten_pentagram()
        print_error_edge_version()
        raise ValueError

    msedgedriver_path = path.join(download_msedgedriver_path, 'msedgedriver.exe')
    
    # 如果已经存在适配当前浏览器版本的 msedgedriver.exe，则无需下载
    if path.exists(msedgedriver_path):
        print_info_edge_existed(version, msedgedriver_path)
        return msedgedriver_path

    # 使用新的微软官方CDN域名，并提供旧域名作为备选
    urls = [
        f"https://msedgedriver.microsoft.com/{version}/edgedriver_win64.zip",
        f"https://msedgedriver.azureedge.net/{version}/edgedriver_win64.zip"
    ]

    # 尝试从多个URL下载
    for url in urls:
        try:
            response = get(url, timeout=30)
            if response.status_code == 200:
                with ZipFile(BytesIO(response.content)) as z:
                    z.extractall(download_msedgedriver_path)
                print_info_edge_download(download_msedgedriver_path)
                return msedgedriver_path
        except Exception as e:
            print(f" --> 尝试从 {url} 下载失败，尝试下一个地址...")
            continue
    
    # 所有URL都失败
    print_ten_pentagram()
    print_error_edge_download(version)
    return None

# 配置 Selenium 的 Edge 驱动
def setup_driver(download_msedgedriver_path, download_dir):
    
    print_info_edge_preparing('配置')

    edge_options = Options()
    
    prefs = {
    "download.default_directory": download_dir,  # 设置下载目录
    "download.prompt_for_download": False,  # 禁止下载提示
    "download.directory_upgrade": True,  # 允许覆盖下载目录
    "profile.default_content_setting_values.automatic_downloads": 1,  # 允许自动下载
    "profile.default_content_setting_values.popups": 0,  # 禁用弹出窗口
    "profile.default_content_setting_values.mixed_script": 1  # 允许混合内容
    }
    
    edge_options.add_experimental_option("prefs", prefs)
    
    # 下载【msedgedriver.exe】的路径
    msedgedriver_path = path.join(download_msedgedriver_path, 'msedgedriver.exe')

    service = EdgeService(executable_path = msedgedriver_path)
    driver = Edge(service = service, options = edge_options)
    
    return driver

# 遍历文件夹并找到大【.stl】文件
def find_large_stl_files(meshes_folder_path, size_threshold_mb):
    large_stl_files_path = []
    for file_name in listdir(meshes_folder_path):
        if file_name.endswith('.stl'):
            file_path = path.join(meshes_folder_path, file_name)
            if path.getsize(file_path) > size_threshold_mb * 1024 * 1024:
                large_stl_files_path.append(file_path)
    return large_stl_files_path

# 获取滑动条的当前位置
def get_current_slider_position(driver, compression_slider):
    current_value = driver.execute_script("return arguments[0].value;", compression_slider)
    return int(current_value)

# 根据文件大小和滑动条当前位置计算调整量
def calculate_compression_adjustment(file_size, current_position, size_threshold_mb):

    # 阈值大小
    size_threshold_bytes = size_threshold_mb * 1024 * 1024

    needed_compression_percentage = (100 * size_threshold_bytes / file_size) - current_position

    # 向下取整，避免超出阈值
    return floor(needed_compression_percentage)

# 调整滑动条
def adjust_compression_slider(driver, compression_slider, file_size, size_threshold_mb):

    # 获取当前滑动条位置
    current_position = get_current_slider_position(driver, compression_slider)

    # 计算调整量
    adjustment = calculate_compression_adjustment(file_size, current_position, size_threshold_mb)
    
    # 如果不需要调整，直接返回
    if adjustment == 0:
        return

    actions = ActionChains(driver)
    
    if adjustment > 0:  # 调整量大于0，目标位置在右侧
        actions.click(compression_slider).send_keys(Keys.ARROW_RIGHT * adjustment).perform()
    else:  # 调整量小于0，目标位置在左侧
        adjustment = abs(adjustment)
        actions.click(compression_slider).send_keys(Keys.ARROW_LEFT * adjustment).perform()

    # 确保调整完成后稍等片刻
    sleep(0.5)

# 等待上传按钮、滑动条、简化按钮、下载按钮都加载完成
def wait_for_upload_and_download_buttons(driver):

    upload_condition = presence_of_element_located((By.XPATH, '//input[@type="file"]'))
    slider_condition = presence_of_element_located((By.XPATH, '//input[@type="range"]'))
    simplify_condition = presence_of_element_located((By.XPATH, '//button[@onclick="uploaded()"]'))
    download_condition = presence_of_element_located((By.ID, 'download'))
    
    WebDriverWait(driver, 10).until(lambda d: upload_condition(driver) and slider_condition(driver) and simplify_condition(driver) and download_condition(driver))

# 检测类名变化
def wait_for_class_change(driver):
    download_condition = presence_of_element_located((By.ID, 'download'))
    
    WebDriverWait(driver, 10).until(
        lambda d: "download_ready_signal" in driver.find_element(By.ID, 'download').get_attribute("class")
    )

# 获取下载目录最新的文件
def get_latest_downloaded_file(meshes_folder_path, timeout=30, interval=1):

    # 检测下载目录，并返回最新下载的文件路径
    # meshes_folder_path: 下载目录路径
    # timeout: 超时时间，秒
    # interval: 检测间隔，秒
    # return: 最新的文件路径

    end_time = time() + timeout
    while True:

        # 获取目录下的所有文件，按修改时间排序
        files = [path.join(meshes_folder_path, f) for f in listdir(meshes_folder_path)]
        files = [f for f in files if path.isfile(f)]
        
        if files:

            # 根据文件的最后修改时间排序，获取最新的文件
            latest_file = max(files, key=path.getmtime)

            # 如果文件已经存在并且是完整文件，返回文件路径
            if latest_file.endswith('.stl') and not latest_file.endswith('.crdownload'):
                return latest_file
        
        if time() > end_time:
            break
        sleep(interval)
    return None

# 上传文件，调整压缩比并下载
def upload_and_simplify_stl(driver, file_path, meshes_folder_path, size_threshold_mb):

    # 获取文件大小
    file_size = path.getsize(file_path)

    # 将文件路径转换为绝对路径
    absolute_file_path = path.abspath(file_path)

    # 访问 STL 简化网站
    driver.get('https://myminifactory.github.io/Fast-Quadric-Mesh-Simplification/')
    driver.minimize_window()  # 最小化
    
    # 等待上传按钮、滑动条、简化按钮、下载按钮都加载完成
    wait_for_upload_and_download_buttons(driver)

    # 获取上传按钮、滑动条、简化按钮、下载按钮
    upload_input = driver.find_element(By.XPATH, '//input[@type="file"]')
    compression_slider = driver.find_element(By.XPATH, '//input[@type="range"]')
    simplify_button = driver.find_element(By.XPATH, '//button[@onclick="uploaded()"]')
    download_button = driver.find_element(By.XPATH, '//button[@id="download"]')

    # 调整滑动条
    adjust_compression_slider(driver, compression_slider, file_size, size_threshold_mb)

    # 上传文件
    upload_input.send_keys(absolute_file_path)
    print_info_file_uploading(absolute_file_path)
    
    # 等待文件上传完成（检测下载按钮的类名变化）
    wait_for_class_change(driver)

    # 点击简化按钮
    simplify_button.click()

    # 等待文件简化完成
    sleep(1)

    # 点击下载按钮
    download_button.click()

    # 等待文件下载完成
    sleep(2)

    # 等待下载完成，获取最新下载的文件路径
    downloaded_file_path = get_latest_downloaded_file(meshes_folder_path)
    print_info_file_download(downloaded_file_path)

# 将 meshes 文件夹中的多余文件删除，并删除无前缀的文件，重命名带前缀的文件
def modify_mesh_folder_files_again(work_space, urdf_file):

    # 找到 meshes 文件夹
    meshes_folder_path = path.join(work_space, 'meshes')
    # 检查 meshes 文件夹是否存在
    if not path.exists(meshes_folder_path):
        print_ten_pentagram()
        print_error_meshes_stl()
        raise FileNotFoundError

    # 获取文件夹中所有文件的名称
    mesh_files = listdir(meshes_folder_path)

    # 解析 URDF 文件并获取所有连杆名称
    tree = parse(urdf_file)
    root = tree.getroot()
    link_names = [link.get('name') for link in root.findall('.//link')]

    # 遍历 URDF 中的每个连杆名称，检查对应的【.STL】文件
    for link_name in link_names:
        stl_file = f"{link_name}.STL"
        stl_file_lower = f"{link_name}.stl"

        # 如果连杆的【.STL】文件既不以大写也不以小写存在，则报错
        if stl_file not in mesh_files and stl_file_lower not in mesh_files:
            print_ten_pentagram()
            print_error_meshes_stl(link_name)
            raise FileNotFoundError

        # 如果【.STL】文件存在，且是大写，则修改为小写
        if stl_file in mesh_files:
            new_file_name = stl_file.replace('.STL', '.stl')
            rename(path.join(meshes_folder_path, stl_file), path.join(meshes_folder_path, new_file_name))
            print_info_meshes_stl('重命名', stl_file, new_file_name)

        # 删除多余文件（以(n)结尾为特征，误操作造成）
        for file_name in mesh_files:
            if file_name.endswith('.stl'):
                if search(r'\(\d+\)\.stl$', file_name):
                    file_path = path.join(meshes_folder_path, file_name)

                    # 在删除文件前检查是否存在，避免找不到文件的错误
                    if path.exists(file_path):
                        remove(file_path)
                        print_info_meshes_stl('删除', file_name)

        # 检查是否有带前缀的文件
        prefixed_file = f"simplify_{link_name}.stl"
        if prefixed_file in mesh_files:
            original_file_path = path.join(meshes_folder_path, stl_file_lower)
            prefixed_file_path = path.join(meshes_folder_path, prefixed_file)
            
            # 删除无前缀的文件，重命名带前缀的文件
            if path.exists(original_file_path):
                remove(original_file_path)
                print_info_meshes_stl('删除', stl_file_lower)
                
            rename(prefixed_file_path, original_file_path)
            print_info_meshes_stl('重命名', prefixed_file, stl_file_lower)

# 第二个功能小函数结束

# 启动浏览器，准备文件
# get_edge_version()
# download_msedgedriver(download_msedgedriver_path)
# setup_driver(download_msedgedriver_path, download_dir)
# find_large_stl_files(meshes_folder_path, size_threshold_mb)

# 上传文件，下载文件
# get_current_slider_position(driver, compression_slider)
# calculate_compression_adjustment(file_size, current_position, size_threshold_mb)
# adjust_compression_slider(driver, compression_slider, file_size, size_threshold_mb)
# wait_for_upload_and_download_buttons(driver)
# wait_for_class_change(driver)
# get_latest_downloaded_file(meshes_folder_path, timeout=30, interval=1)
# upload_and_simplify_stl(driver, file_path, meshes_folder_path, size_threshold_mb)

# 文件重命名
# modify_mesh_folder_files_again(work_space, urdf_file)
# find_files(work_space)  # 第〇个功能小函数

# 第二个功能大函数 mesh_simplify 开始
def mesh_simplify(work_space, size_threshold_mb, download_msedgedriver_path, initial_excel_file):

    if not isinstance(work_space, str) or not path.isdir(work_space):
        print_ten_pentagram()
        print_error_void('工作目录')
        return
    
    if not isinstance(size_threshold_mb, (int, float)):
        print_ten_pentagram()
        print_error_void('网格文件简化目标大小')
        return
    
    try:
        urdf_file = find_urdf_file(work_space)
        print_info_file_found(urdf_file, 'URDF')
    except Exception as e:
        print_ten_pentagram()
        print_error_file_not_found('URDF')
        return
    
    try:
        download_msedgedriver(download_msedgedriver_path)
    except Exception as e:
        print_ten_pentagram()
        print_error_edge_download_details(e)
        return

    meshes_folder_path = path.join(work_space, 'meshes')
    if not path.exists(meshes_folder_path):
        print_ten_pentagram()
        print_error_meshes_stl()
        return
    
    # 设置 Selenium 驱动
    try:
        driver = setup_driver(download_msedgedriver_path, meshes_folder_path)
        if not driver:
            print_ten_pentagram()
            print_error_edge_initialization()
            return
    except Exception as e:
        print_ten_pentagram()
        print_error_edge_initialization()
        return

    try:
        while True:
            # 找到大于阈值大小的【.stl】文件
            large_stl_files = find_large_stl_files(meshes_folder_path, size_threshold_mb)

            if not large_stl_files:
                print_info_size_detection(size_threshold_mb)
                break

            # 遍历每个大文件，进行简化
            for large_stl_file in large_stl_files:
                upload_and_simplify_stl(driver, large_stl_file, meshes_folder_path, size_threshold_mb)

            # 将 meshes 文件夹中的【.STL】扩展名修改为【.stl】
            modify_mesh_folder_files_again(work_space, urdf_file)

        print_info_task2()
    except Exception as e:
        print_ten_pentagram()
        print_error_task2()
        return
    
    finally:
        # 关闭浏览器
        driver.quit()
# 第二个功能大函数 mesh_simplify 结束

# 第一个功能大函数
# urdf_data_replace(work_space, size_threshold_mb, download_msedgedriver_path, initial_excel_file)
# 第二个功能大函数
# mesh_simplify(work_space, size_threshold_mb, download_msedgedriver_path, initial_excel_file)

# 功能整合
def only_urdf_data_replace(work_space, size_threshold_mb, download_msedgedriver_path, initial_excel_file):
    urdf_data_replace(work_space, size_threshold_mb, download_msedgedriver_path, initial_excel_file)
    print_info_fast_urdf_finish()

def urdf_data_replace_and_mesh_simplify(work_space, size_threshold_mb, download_msedgedriver_path, initial_excel_file):
    urdf_data_replace(work_space, size_threshold_mb, download_msedgedriver_path, initial_excel_file)
    mesh_simplify(work_space, size_threshold_mb, download_msedgedriver_path, initial_excel_file)
    print_info_fast_urdf_finish()

# PrintStream 类用于重定向标准输出（print）到 GUI 文本框，继承自 QObject
class PrintStream(QObject):

    # 定义信号 signal，用于发送字符串消息
    signal = pyqtSignal(str)

    # write 方法接收消息并通过信号发送给 GUI
    def write(self, message):
        self.signal.emit(message)

# MainWorker 类用于在单独的线程中运行指定的功能，继承自 QThread
class MainWorker(QThread):
    output_signal = pyqtSignal(str)

    # 构造函数初始化所需参数
    def __init__(self, func, work_space, size_threshold_mb, download_msedgedriver_path, initial_excel_file):
        super().__init__()

        # func 是要执行的功能
        self.func = func

        # work_space、size_threshold_mb 和 download_msedgedriver_path 是传递给该功能的参数
        self.work_space = work_space
        self.size_threshold_mb = size_threshold_mb
        self.download_msedgedriver_path = download_msedgedriver_path
        self.initial_excel_file = initial_excel_file

    def run(self):
        self.func(self.work_space, self.size_threshold_mb, self.download_msedgedriver_path, self.initial_excel_file)

class MainWindow(QWidget):
    def __init__(self):
        super().__init__()

        # 初始化线程变量
        self.worker = None

        self.setStyleSheet("""
            QWidget {
                font-size: 16px;
                color: #333333;
                font-family: 'Microsoft YaHei';
                background-color: #FDE6E0;
            }
                  
            QTextEdit {
                border: none;
            }
                           
            QPushButton {
                color: #666666;
                height: 30px;
            }

            QPushButton:hover {
                color: #333333;
            }

            QLineEdit{
                background-color: #FFFFFF;
                height: 30px;
            }

            QTextEdit#step1 {
                color: #333333;
                font-size: 12px;
            }
                           
            QTextEdit#step2 {
                margin-top: 10px;
                border-top: 1px solid #333333;
                color: #333333;
                font-size: 12px;
            }

            QTextEdit#step3 {
                margin-top: 10px;
                border-top: 1px solid #333333;
                color: #333333;
                font-size: 12px;
            }

            QTextEdit#step4 {
                margin-top: 10px;
                border-top: 1px solid #333333;
                color: #333333;
                font-size: 12px;
            }
                           
            QTextEdit#finish {
                margin-top: 10px;
                border-top: 1px solid #333333;
                color: #333333;
                font-size: 12px;
            }
                           
            QPushButton#initial_excel_file_button_run {
                margin: 0px 10px 0px 10px;
                background-color: #E1E3FE;
                border: 3px solid #DCE2F1;
                border-top: 3px solid #E9EBFE;
                border-radius: 5px;
                font-size: 20px;
                font-weight: 700;
            }
                           
            QPushButton#initial_excel_file_button_run:hover {
                background-color: #DBDDFA;
            }
                           
            QPushButton#initial_excel_file_button_run:pressed {
                background-color: #D9DBF8;
            }
                   
            QPushButton#floor_button {
                margin-left: 10px;
                background-color: #C7EDCC;
                border: 3px solid #CCE8CF;
                border-top: 3px solid #CFF4C3;
                border-radius: 5px;
                font-size: 30px;
                font-weight: 700;
            }
                           
            QPushButton#floor_button:hover {
                background-color: #C1E7C6;
            }
                           
            QPushButton#floor_button:pressed {
                background-color: #BEE4C3;
            }

            QPushButton#ceil_button {
                margin-left: 10px;
                background-color: #C7EDCC;
                border: 3px solid #CCE8CF;
                border-top: 3px solid #CFF4C3;
                border-radius: 5px;
                font-size: 30px;
                font-weight: 700;
            }
                           
            QPushButton#ceil_button:hover {
                background-color: #C1E7C6;
            }
                           
            QPushButton#ceil_button:pressed {
                background-color: #BEE4C3;
            }
                            
            QPushButton#button1 {
                margin: 10px;
                margin-top: 0;
                height: 80px;
                background-color: #E1E3FE;
                border: 3px solid #DCE2F1;
                border-top: 3px solid #E9EBFE;
                border-radius: 15px;
                font-size: 25px;
            }
                           
            QPushButton#button1:hover {
                background-color: #DBDDFA;
            }
                           
            QPushButton#button1:pressed {
                background-color: #D9DBF8;
            }
                           
            QPushButton#button2 {
                margin: 10px;
                margin-top: 0;
                height: 80px;
                background-color: #E1E3FE;
                border: 3px solid #DCE2F1;
                border-top: 3px solid #E9EBFE;
                border-radius: 15px;
                font-size: 25px;
            }
                           
            QPushButton#button2:hover {
                background-color: #DBDDFA;
            }
                           
            QPushButton#button2:pressed {
                background-color: #D9DBF8;
            }
                                                      
            QTextEdit#main_text {
                margin-bottom: 10px;
                font-size: 15px;
                color: #333333;
                border-radius: 15px;
                background-color: #FFFFFF;
            }
                           
            QTextEdit#dividing_line {
                border-top: 1px solid #555555;
                color: #555555;
                width: 100%;
                font-size: 12px;
                line-height: 12px;
            }
        """)

        self.setWindowTitle(f'Fast URDF {__version__} -- by Mxqwthl')  # 主窗口名称
        self.setGeometry(100, 100, 900, 556)  # 主窗口位置和大小

        # 设置图标路径
        if getattr(sys, 'frozen', False):
            icon_path = path.join(sys._MEIPASS, 'favicon.ico')
        else:
            icon_path = 'favicon.ico'
        self.setWindowIcon(QIcon(icon_path))

        # 设置窗口的最小宽度和高度
        self.setMinimumSize(550, 556)

        # 主布局（纵向）
        self.layout = QVBoxLayout()
        self.layout.setSpacing(0)
        self.layout.setContentsMargins(10, 6, 10, 4)

        # Step1
        self.step1 = QTextEdit(self)
        self.step1.setObjectName('step1')
        self.step1.setHtml('<p><strong>Step1：</strong>选择工作目录</p>')
        self.step1.setReadOnly(True)  # 设置为只读
        self.step1.setFixedHeight(26)
        self.layout.addWidget(self.step1)

        # 输入 work_space 副布局（横向）
        self.work_space_layout = QHBoxLayout()
        # self.work_space_layout.setSpacing(0)
        # self.work_space_layout.setContentsMargins(0, 0, 0, 0)
        self.layout.addLayout(self.work_space_layout)
        # 输入 work_space
        self.work_space_input = QLineEdit(self)
        self.work_space_input.setObjectName('work_space_input')
        self.work_space_input.setPlaceholderText('<导出 URDF 时创建的文件夹>')
        self.work_space_layout.addWidget(self.work_space_input)
        # 输入 work_space 按钮
        self.work_space_button = QPushButton('选择文件夹', self)
        self.work_space_button.setObjectName('work_space_button')
        self.work_space_button.clicked.connect(self.work_space_select)
        self.work_space_layout.addWidget(self.work_space_button)

        self.work_space_input.setMinimumWidth(0)  # 允许扩展
        self.work_space_input.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        self.work_space_button.setMinimumWidth(0)  # 允许扩展
        self.work_space_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)

        self.work_space_layout.setStretch(0, 4)
        self.work_space_layout.setStretch(1, 1)

        # Step2
        self.step2 = QTextEdit(self)
        self.step2.setObjectName('step2')
        self.step2.setHtml('<p><strong>Step2：</strong>选择原始 Excel 文件，点击生成【若已生成过或手动调整好，请忽略此步骤】</p>')
        self.step2.setReadOnly(True)  # 设置为只读
        self.step2.setFixedHeight(36)
        self.layout.addWidget(self.step2)

        # 输入 initial_excel_file 副布局（横向）
        self.initial_excel_file_layout = QHBoxLayout()
        self.initial_excel_file_layout.setSpacing(0)
        self.initial_excel_file_layout.setContentsMargins(0, 0, 0, 0)
        self.layout.addLayout(self.initial_excel_file_layout)
        # 输入 initial_excel_file
        self.initial_excel_file_input = QLineEdit(self)
        self.initial_excel_file_input.setObjectName('initial_excel_file_input')
        self.initial_excel_file_input.setPlaceholderText('<从 Solidworks 中粘贴数据的 Excel 文件>')
        self.initial_excel_file_layout.addWidget(self.initial_excel_file_input)
        # 输入 initial_excel_file 按钮
        self.initial_excel_file_button = QPushButton('选择文件', self)
        self.initial_excel_file_button.setObjectName('initial_excel_file_button')
        self.initial_excel_file_button.clicked.connect(self.initial_excel_file_select)
        self.initial_excel_file_layout.addWidget(self.initial_excel_file_button)
        # 处理 initial_excel_file 按钮
        self.initial_excel_file_button_run = QPushButton('生成', self)
        self.initial_excel_file_button_run.setObjectName('initial_excel_file_button_run')
        self.initial_excel_file_button_run.clicked.connect(self.run_build_excel_file)
        self.initial_excel_file_layout.addWidget(self.initial_excel_file_button_run)

        self.initial_excel_file_input.setMinimumWidth(0)  # 允许扩展
        self.initial_excel_file_input.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        self.initial_excel_file_button.setMinimumWidth(0)  # 允许扩展
        self.initial_excel_file_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        self.initial_excel_file_button_run.setMinimumWidth(0)  # 允许扩展
        self.initial_excel_file_button_run.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)

        self.initial_excel_file_layout.setStretch(0, 3)
        self.initial_excel_file_layout.setStretch(1, 1)
        self.initial_excel_file_layout.setStretch(2, 1)

        # Step3
        self.step3 = QTextEdit(self)
        self.step3.setObjectName('step3')
        self.step3.setHtml('<p><strong>Step3：</strong>设定网格文件简化的目标大小</p>')
        self.step3.setReadOnly(True)  # 设置为只读
        self.step3.setFixedHeight(36)
        self.layout.addWidget(self.step3)

        # 输入 size_threshold_mb 副布局（横向）
        self.size_threshold_mb_layout = QHBoxLayout()
        self.size_threshold_mb_layout.setSpacing(0)
        self.size_threshold_mb_layout.setContentsMargins(0, 0, 0, 0)
        self.layout.addLayout(self.size_threshold_mb_layout)
        # 输入 size_threshold_mb
        self.size_threshold_mb_input = QLineEdit(self)
        self.size_threshold_mb_input.setObjectName('size_threshold_mb_input')
        self.size_threshold_mb_input.setPlaceholderText('<请输入数字>')
        self.size_threshold_mb_input.setText('5')  # 设置默认值为5
        self.size_threshold_mb_layout.addWidget(self.size_threshold_mb_input)
        # 输入 size_threshold_mb 单位
        self.size_threshold_mb_input_unit = QTextEdit(self)
        self.size_threshold_mb_input_unit.setObjectName('size_threshold_mb_input_unit')
        self.size_threshold_mb_input_unit.setPlainText('MB')
        self.size_threshold_mb_input_unit.setReadOnly(True)  # 设置为只读
        self.size_threshold_mb_input_unit.setFixedHeight(30)
        self.size_threshold_mb_layout.addWidget(self.size_threshold_mb_input_unit)
        # 输入 size_threshold_mb 增加
        self.floor_button = QPushButton('+', self)
        self.floor_button.setObjectName('floor_button')
        self.floor_button.clicked.connect(self.size_threshold_mb_up)
        self.size_threshold_mb_layout.addWidget(self.floor_button)
        # 输入 size_threshold_mb 减少
        self.ceil_button = QPushButton('\u2013', self)
        self.ceil_button.setObjectName('ceil_button')
        self.ceil_button.clicked.connect(self.size_threshold_mb_down)
        self.size_threshold_mb_layout.addWidget(self.ceil_button)
        # 输入 size_threshold_mb 占位
        self.size_threshold_mb_input_block = QTextEdit(self)
        self.size_threshold_mb_input_block.setObjectName('size_threshold_mb_input_block')
        self.size_threshold_mb_input_block.setPlainText('')
        self.size_threshold_mb_input_block.setReadOnly(True)  # 设置为只读
        self.size_threshold_mb_input_block.setFixedHeight(30)
        self.size_threshold_mb_layout.addWidget(self.size_threshold_mb_input_block)

        self.size_threshold_mb_input.setMinimumWidth(0)  # 允许扩展
        self.size_threshold_mb_input.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        self.size_threshold_mb_input_unit.setMinimumWidth(0)  # 允许扩展
        self.size_threshold_mb_input_unit.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        self.floor_button.setMinimumWidth(0)  # 允许扩展
        self.floor_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        self.ceil_button.setMinimumWidth(0)  # 允许扩展
        self.ceil_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        self.size_threshold_mb_input_block.setMinimumWidth(0)  # 允许扩展
        self.size_threshold_mb_input_block.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)

        self.size_threshold_mb_layout.setStretch(0, 3)
        self.size_threshold_mb_layout.setStretch(1, 1)
        self.size_threshold_mb_layout.setStretch(2, 1)
        self.size_threshold_mb_layout.setStretch(3, 1)
        self.size_threshold_mb_layout.setStretch(4, 6)

        # Step4
        self.step4 = QTextEdit(self)
        self.step4.setObjectName('step4')
        self.step4.setHtml('<p><strong>Step4：</strong>选择 Edge 驱动下载目标路径【不用管】</p>')
        self.step4.setReadOnly(True)  # 设置为只读
        self.step4.setFixedHeight(36)
        self.layout.addWidget(self.step4)

        # 输入 download_msedgedriver_path 副布局（横向）
        self.download_msedgedriver_path_layout = QHBoxLayout()
        self.download_msedgedriver_path_layout.setSpacing(0)
        self.download_msedgedriver_path_layout.setContentsMargins(0, 0, 0, 0)
        self.layout.addLayout(self.download_msedgedriver_path_layout)
        # 输入 download_msedgedriver_path
        self.download_msedgedriver_path_input = QLineEdit(self)
        self.download_msedgedriver_path_input.setObjectName('download_msedgedriver_path_input')
        self.download_msedgedriver_path_input.setPlaceholderText('<默认 C:\\WebDriver>')
        self.download_msedgedriver_path_input.setText('C:\\WebDriver')  # 设置默认值为 C:\\WebDriver
        self.download_msedgedriver_path_layout.addWidget(self.download_msedgedriver_path_input)
        # 输入 download_msedgedriver_path 按钮
        self.download_msedgedriver_path_button = QPushButton('选择文件夹', self)
        self.download_msedgedriver_path_button.setObjectName('download_msedgedriver_path_button')
        self.download_msedgedriver_path_button.clicked.connect(self.download_msedgedriver_path_select)
        self.download_msedgedriver_path_layout.addWidget(self.download_msedgedriver_path_button)
        # # 隐藏
        # self.step4.setVisible(False)
        # self.download_msedgedriver_path_input.setVisible(False)
        # self.download_msedgedriver_path_button.setVisible(False)

        self.download_msedgedriver_path_input.setMinimumWidth(0)  # 允许扩展
        self.download_msedgedriver_path_input.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        self.download_msedgedriver_path_button.setMinimumWidth(0)  # 允许扩展
        self.download_msedgedriver_path_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)

        self.download_msedgedriver_path_layout.setStretch(0, 4)
        self.download_msedgedriver_path_layout.setStretch(1, 1)

        # Step4
        self.finish = QTextEdit(self)
        self.finish.setObjectName('finish')
        self.finish.setHtml('<p><strong>Finish！</strong>请点击按钮【对于 Mujoco 等对网格文件有要求的仿真软件请选择后者】</p>')
        self.finish.setReadOnly(True)  # 设置为只读
        self.finish.setFixedHeight(36)
        self.layout.addWidget(self.finish)

        # 按钮副布局（横向）
        self.buttons_layout = QHBoxLayout()
        self.buttons_layout.setSpacing(0)
        self.buttons_layout.setContentsMargins(0, 0, 0, 0)
        self.layout.addLayout(self.buttons_layout)
        # 按钮1
        self.button1 = QPushButton('Fast URDF\n数据替换', self)
        self.button1.setObjectName('button1')
        self.button1.clicked.connect(self.run_urdf_data_replace)
        self.buttons_layout.addWidget(self.button1)
        # 按钮2
        self.button2 = QPushButton('Fast URDF\n数据替换 + 网格简化', self)
        self.button2.setObjectName('button2')
        self.button2.clicked.connect(self.run_urdf_data_replace_and_mesh_simplify)
        self.buttons_layout.addWidget(self.button2)

        self.buttons_layout.setStretch(0, 1)
        self.buttons_layout.setStretch(1, 1)

        # 主文本框
        self.main_text = QTextEdit(self)
        self.main_text.setObjectName('main_text')
        self.main_text.setReadOnly(True)
        self.layout.addWidget(self.main_text)

        # 分割线
        self.dividing_line = QTextEdit(self)
        self.dividing_line.setObjectName('dividing_line')
        self.dividing_line.setPlainText(' * 网格简化使用【https://myminifactory.github.io/Fast-Quadric-Mesh-Simplification】')
        self.dividing_line.setReadOnly(True)  # 设置为只读
        self.dividing_line.setFixedHeight(26)
        self.layout.addWidget(self.dividing_line)

        self.setLayout(self.layout)

        # 重定向 print 输出
        self.print_stream = PrintStream()
        self.print_stream.signal.connect(self.update_output)
        sys.stdout = self.print_stream
    
    def closeEvent(self, event):
        if self.worker is not None and self.worker.isRunning():
            self.worker.quit()  # 请求线程退出
            self.worker.wait()  # 等待线程完成
        event.accept()  # 继续关闭窗口

    def work_space_select(self):
        work_space_selection = QFileDialog.getExistingDirectory(self, '选择目标文件夹')
        work_space_selection = work_space_selection.replace('/', '\\')
        if work_space_selection:
            self.work_space_input.setText(work_space_selection)

    def initial_excel_file_select(self):
        initial_excel_file_selection, _ = QFileDialog.getOpenFileName(self, '选择目标文件')
        initial_excel_file_selection = initial_excel_file_selection.replace('/', '\\')
        if initial_excel_file_selection:
            self.initial_excel_file_input.setText(initial_excel_file_selection)

    def download_msedgedriver_path_select(self):
        download_msedgedriver_path_selection = QFileDialog.getExistingDirectory(self, '选择目标文件夹')
        download_msedgedriver_path_selection = download_msedgedriver_path_selection.replace('/', '\\')
        if download_msedgedriver_path_selection:
            self.download_msedgedriver_path_input.setText(download_msedgedriver_path_selection)

    def size_threshold_mb_up(self):
        try:
            current_value = float(self.size_threshold_mb_input.text())
            if current_value > 0:
                new_value = floor(current_value) + 1
                self.size_threshold_mb_input.setText(str(new_value))
            else:
                raise ValueError
        except ValueError:
            self.size_threshold_mb_input.setText('5')  # 重置为默认值
            self.main_text.append(' --> 网格文件简化目标大小输入不合法！已重置为默认值：5MB')

    def size_threshold_mb_down(self):
        try:
            current_value = float(self.size_threshold_mb_input.text())
            if current_value <= 0:
                raise ValueError
            if current_value > 1:
                new_value = ceil(current_value) - 1
                self.size_threshold_mb_input.setText(str(new_value))
            else:
                return
        except ValueError:
            self.size_threshold_mb_input.setText('5')  # 重置为默认值
            self.main_text.append(' --> 网格文件简化目标大小输入不合法！已重置为默认值：5MB')

    def run_build_excel_file(self):
        work_space = self.work_space_input.text()
        initial_excel_file = self.initial_excel_file_input.text()
        size_threshold_mb = 5  # 不读取输入框内容，防止闪退     
        download_msedgedriver_path = self.download_msedgedriver_path_input.text()
        self.start_worker(build_excel_file, work_space, size_threshold_mb, download_msedgedriver_path, initial_excel_file)

    def run_urdf_data_replace(self):
        work_space = self.work_space_input.text()
        initial_excel_file = self.initial_excel_file_input.text()
        size_threshold_mb = 5  # 不读取输入框内容，防止闪退     
        download_msedgedriver_path = self.download_msedgedriver_path_input.text()
        self.start_worker(only_urdf_data_replace, work_space, size_threshold_mb, download_msedgedriver_path, initial_excel_file)

    def run_urdf_data_replace_and_mesh_simplify(self):
        try:
            work_space = self.work_space_input.text()
            initial_excel_file = self.initial_excel_file_input.text()
            size_threshold_mb = float(self.size_threshold_mb_input.text())

            if size_threshold_mb:
                if size_threshold_mb <= 0:
                    raise ValueError
            else:
                raise ValueError
            
            download_msedgedriver_path = self.download_msedgedriver_path_input.text()
            self.start_worker(urdf_data_replace_and_mesh_simplify, work_space, size_threshold_mb, download_msedgedriver_path, initial_excel_file)

        except ValueError:
            self.size_threshold_mb_input.setText('5')  # 重置为默认值
            self.main_text.append(' --> 网格文件简化目标大小输入不合法！已重置为默认值：5MB')

    def start_worker(self, func, work_space, size_threshold_mb, download_msedgedriver_path, initial_excel_file):
        self.worker = MainWorker(func, work_space, size_threshold_mb, download_msedgedriver_path, initial_excel_file)
        self.worker.output_signal.connect(self.update_output)
        self.worker.start()

    def update_output(self, output):
        self.main_text.append(output.rstrip())

def main():
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())

if __name__ == '__main__':
    main()

# 代码结构：
# # 第〇个功能大函数
# def build_excel_file(work_space, size_threshold_mb, download_msedgedriver_path, initial_excel_file):
# # 第一个功能大函数
# def urdf_data_replace(work_space, size_threshold_mb, download_msedgedriver_path, initial_excel_file):
# # 第二个功能大函数
# def mesh_simplify(work_space, size_threshold_mb, download_msedgedriver_path, initial_excel_file):
# # 功能整合
# def only_urdf_data_replace(work_space, size_threshold_mb, download_msedgedriver_path, initial_excel_file):
#     urdf_data_replace(work_space, size_threshold_mb, download_msedgedriver_path, initial_excel_file)
# def urdf_data_replace_and_mesh_simplify(work_space, size_threshold_mb, download_msedgedriver_path, initial_excel_file):
#     urdf_data_replace(work_space, size_threshold_mb, download_msedgedriver_path, initial_excel_file)
#     mesh_simplify(work_space, size_threshold_mb, download_msedgedriver_path, initial_excel_file)
# # 重定向标准输出
# class PrintStream(QObject)
# # 主线程
# class MainWorker(QThread)
# # 主窗口
# class MainWindow(QWidget)
# # 主函数
# def main():
#     app = QApplication(sys.argv)
#     window = MainWindow()
#     window.show()
#     sys.exit(app.exec_())
# if __name__ == '__main__':
#     main()